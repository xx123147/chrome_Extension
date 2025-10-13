import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = load_workbook(r'C:\Users\Administrator\Desktop\上新\ASIN\10.7.xlsx')

# wb = Workbook("result.xlsx")
# ws = wb.create_sheet(title='asin')
ws = wb['Sheet3']

URL = "https://www.amazon.co.uk/s"
headers = {
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language":"en-GB,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.amazon.co.uk/",
}
params = {
    "k":"Alaeseje",
    "crid":"F46WM0W90HYZ"
}
num = 1
for page in range(1,3):
    try:
        p = params.copy()
        p["page"] = page
        response = requests.get(URL,headers=headers,params=p)
        tree = etree.HTML(response.text)
        data_asin = tree.xpath("//div[@role='listitem' and not(contains(.,'Sponsored'))]/@data-asin")
        for asin in data_asin:
            ws.cell(row=num, column=1, value=asin)
            num+=1
    except Exception as ex:
        print(f"第{page}页出现问题,{ex}")

        # print(data_asin)


wb.save(r'C:\Users\Administrator\Desktop\上新\ASIN\10.7.xlsx')