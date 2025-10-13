import re
from traceback import print_tb

import openpyxl
from playwright.sync_api import Playwright, sync_playwright, expect


def run(playwright: Playwright,title_list):
    try:
        result=[]
        #创建一次页面环境
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto("https://www.amazon.co.uk/ref=nav_logo")
        page.get_by_role("button", name="Accept").click()
        page.get_by_role("button", name="Submit").nth(1).click()
        page.get_by_role("textbox", name="or enter a UK mainland").fill("BH21 2EY")
        page.get_by_label("Apply").click()
        page.get_by_role("button", name="Continue").click()
        page.wait_for_timeout(3000)

        # 循环处理所有标题，复用同一个页面
        for i,title in enumerate(title_list,1):
            try:
                page.get_by_role("searchbox", name="Search Amazon.co.uk").fill(title)
                page.get_by_role("searchbox", name="Search Amazon.co.uk").press("Enter")
                page.get_by_role("button", name="Go", exact=True).click()
                page.wait_for_timeout(3000)
                # 获得asin
                non_sponsored_item = page.locator("xpath=//div[@role='listitem' and not(contains(.,'Sponsored')) and @data-asin]")
                asin = non_sponsored_item.first.get_attribute("data-asin")
                result.append(asin)
                page.go_back()
            except Exception as e:
                print(f"获取第{i}个asin时发生错误：{e}")
                result.append(None)
                page.goto("https://www.amazon.co.uk/ref=nav_logo")
                page.wait_for_timeout(3000)
    except Exception as ex:
        print(f"初始化出现问题：{ex}")
    finally:
        #统一清理页面
        page.close()
        context.close()
        browser.close()
        print("页面清理完毕！")
        return result



# with sync_playwright() as playwright:
#     run(playwright, "AiliveSun 2 PCS Head Harness Assembly for 6000 Series Half (for6300/6200/6100 Respirator)")


with sync_playwright() as playwright:
    file_path=r'C:\Users\Administrator\Desktop\test.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    title_list=[]
    for row in range(1,ws.max_row+1):
        title = ws.cell(row,1).value
        title_list.append(title)
    # print(title_list)
    result_list = run(playwright,title_list)
    if not result_list:
        print("集合为空！")
    for row in range(1,ws.max_row+1):
        ws.cell(row,3,value=result_list[row-1])


    wb.save(file_path)

