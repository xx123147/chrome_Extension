import re
from traceback import print_tb

import openpyxl
from playwright.sync_api import Playwright, sync_playwright, expect


def run(playwright: Playwright,title) -> str:
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://www.amazon.co.uk/ref=nav_logo")
    # page.get_by_role("button", name="Accept").click()
    # page.get_by_role("button", name="Submit").first.click()
    # page.get_by_role("searchbox", name="Search Amazon.co.uk").click()
    page.get_by_role("searchbox", name="Search Amazon.co.uk").fill(title)
    page.get_by_role("searchbox", name="Search Amazon.co.uk").press("Enter")
    page.get_by_role("button", name="Go", exact=True).click()
    # page.locator(".s-widget-container > span > .puis-card-container > .a-section.a-spacing-base > .s-product-image-container > .rush-component > .a-link-normal").first.click()
    # i=1
    # item=None
    # item_xpath=f"(//div[@role='listitem' and boolean(@data-asin)])[{i}]"
    # while True:
    #     item_xpath = f"(//div[@role='listitem' and boolean(@data-asin)])[{i}]"
    #     item=page.locator(item_xpath)
    #     is_Sponsored=item.locator(":has-text('Sponsored')").count()>0
    #     if is_Sponsored:
    #         print(f"{i} Sponsored,跳过")
    #         i=i+1
    #     else :
    #         break
    #
    # print(item_xpath)
    page.wait_for_timeout(4000)
    non_sponsored_items=page.locator("xpath=//div[@role='listitem' and @data-asin and not(contains(., 'Sponsored'))]//a")
    # print(f"找到 {non_sponsored_items.count()} 个非Sponsored商品")
    src = non_sponsored_items.first.get_attribute("href")
    href = "https://www.amazon.co.uk/"+src
    # non_sponsored_items.first.click(force=True)
    # print(href)
    page.goto(href)
    asin = page.locator("//li[contains(., 'ASIN')]/span/span[2]").inner_text()
    page.close()

    context.close()
    browser.close()
    return asin



with sync_playwright() as playwright:
    file_path=r'C:\Users\Administrator\Desktop\test.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    title=''
    for row in range(1,ws.max_row+1):
        title = ws.cell(row,1).value
        asin=run(playwright,title)
        print(asin)
        ws.cell(row,2,value=asin)
    wb.save(file_path)


