from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

filePath = r'C:\Users\Administrator\Desktop\上新\ASIN\9.30.xlsx'
target = 'Xiaomi'
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
wb = load_workbook(filePath)
ws = wb.active
for row in range(1, ws.max_row + 1):
    c = ws.cell(row, 1)
    if target in c.value:
        ws.cell(row,3,value=1)
        c.fill=yellow_fill
    else:
        ws.cell(row,3,value=0)
wb.save(filePath)
print("检索完成！")