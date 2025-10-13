from openpyxl import load_workbook

file_path = r'C:\Users\Administrator\Desktop\上新\ASIN\9.26\王武俊.xlsx'
try:
    wb = load_workbook(file_path, data_only=True)
    # sheet_names=wb.sheetnames
    # print(sheet_names)
    ws = wb['Sheet4']
    # print(ws.max_row)
    for A_item in range(1, ws.max_row+1):
        v = ws.cell(A_item,1).value
        print(v)
except Exception as e:
    print(e)